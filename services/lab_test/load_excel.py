from pathlib import Path
import openpyxl
from core.database import get_psycopg2_connection as get_connection

EXCEL_PATH = str(Path(__file__).resolve().parent / "Lab test interpretation.xlsx")

PCOS_REASON_MAPPING = {
    "Total Testosterone":                   "High Androgens",
    "Free Testosterone":                    "High Androgens",
    "DHEA-S":                               "High Androgens",
    "Androstenedione":                      "High Androgens",
    "LH:FSH Ratio":                         "High LH:FSH Ratio",
    "Luteinizing Hormone (LH)":             "High LH:FSH Ratio",
    "Follicle-Stimulating Hormone (FSH)":   "High LH:FSH Ratio",
    "Fasting Insulin":                       "High Insulin",
    "Anti-M\u00fcllerian Hormone (AMH)":    "High AMH",

}

PREGNANCY_REASON_MAPPING = {
    "Human Chorionic Gonadotropin (hCG)":  "hCG",
    "Progesterone":                         "Progesterone",
    "Estrogen":                             "Estrogen",
    "Human Placental Lactogen (hPL)":       "hPL",
    "Prolactin":                            "Prolactin",
}


def parse_excel():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    pcos_base = {}
    for row in ws.iter_rows(min_row=5, max_row=14, min_col=1, max_col=3, values_only=True):
        test_name, normal_range, pcos_indication = row
        if test_name:
            pcos_base[test_name.strip()] = {
                "normal_range": normal_range,
                "pcos_indication": pcos_indication,
            }

    pcos_flags = {}
    for row in ws.iter_rows(min_row=5, max_row=7, min_col=5, max_col=7, values_only=True):
        test_name, flag_threshold, flag_condition = row
        if test_name:
            pcos_flags[test_name.strip()] = {
                "flag_threshold": flag_threshold,
                "flag_condition": flag_condition,
            }

    pregnancy_roles = {}
    for row in ws.iter_rows(min_row=20, max_row=24, min_col=1, max_col=2, values_only=True):
        hormone_name, pregnancy_role = row
        if hormone_name:
            pregnancy_roles[hormone_name.strip()] = pregnancy_role

    pregnancy_flags = {}
    for row in ws.iter_rows(min_row=26, max_row=29, min_col=1, max_col=3, values_only=True):
        feature, flag_value, flag_condition = row
        if feature and flag_value:
            name = feature.strip().replace(" Level", "").strip()
            pregnancy_flags[name] = {
                "flag_value": str(flag_value).strip(),
                "flag_condition": flag_condition,
            }

    pcos_reasons = {}
    reason_keys = ["High Androgens", "High LH:FSH Ratio", "High Insulin", "High AMH"]
    for i, row_idx in enumerate(range(35, 39)):
        cell_val = ws.cell(row=row_idx, column=2).value
        if cell_val:
            pcos_reasons[reason_keys[i]] = cell_val.strip()

    pregnancy_reasons = {}
    preg_keys = ["hCG", "Progesterone", "Estrogen", "hPL", "Prolactin"]
    for i, row_idx in enumerate(range(42, 47)):
        cell_val = ws.cell(row=row_idx, column=2).value
        if cell_val:
            pregnancy_reasons[preg_keys[i]] = cell_val.strip()

    return pcos_base, pcos_flags, pregnancy_roles, pregnancy_flags, pcos_reasons, pregnancy_reasons


def load():
    pcos_base, pcos_flags, pregnancy_roles, pregnancy_flags, pcos_reasons, pregnancy_reasons = parse_excel()

    conn = get_connection()
    cur = conn.cursor()

    cur.execute("TRUNCATE TABLE pcos_tests RESTART IDENTITY CASCADE;")
    cur.execute("TRUNCATE TABLE pregnancy_tests RESTART IDENTITY CASCADE;")

    pcos_insert = (
        "INSERT INTO pcos_tests "
        "(test_name, normal_range, pcos_indication, flag_threshold, flag_condition, abnormal_reason) "
        "VALUES (%s, %s, %s, %s, %s, %s)"
    )

    for test_name, base in pcos_base.items():
        flags = pcos_flags.get(test_name, {})
        reason_key = PCOS_REASON_MAPPING.get(test_name)
        abnormal_reason = pcos_reasons.get(reason_key) if reason_key else None
        cur.execute(pcos_insert, (
            test_name,
            base.get("normal_range"),
            base.get("pcos_indication"),
            flags.get("flag_threshold"),
            flags.get("flag_condition", "PCOS"),
            abnormal_reason,
        ))

    flag_name_map = {
        "hCG":          "Human Chorionic Gonadotropin (hCG)",
        "Progesterone": "Progesterone",
        "Estrogen":     "Estrogen",
    }

    preg_insert = (
        "INSERT INTO pregnancy_tests "
        "(hormone_name, pregnancy_role, flag_value, flag_condition, reason_for_change) "
        "VALUES (%s, %s, %s, %s, %s)"
    )

    for hormone_name, role in pregnancy_roles.items():
        short_key = next((k for k, v in flag_name_map.items() if v == hormone_name), None)
        flag_data = pregnancy_flags.get(short_key, {}) if short_key else {}
        reason_key = PREGNANCY_REASON_MAPPING.get(hormone_name)
        reason_for_change = pregnancy_reasons.get(reason_key) if reason_key else None
        cur.execute(preg_insert, (
            hormone_name,
            role,
            flag_data.get("flag_value"),
            flag_data.get("flag_condition", "Pregnancy"),
            reason_for_change,
        ))

    conn.commit()
    cur.close()
    conn.close()
    print("All rows loaded successfully.")


if __name__ == "__main__":
    load()